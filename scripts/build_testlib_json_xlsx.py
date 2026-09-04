#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build tests.json directly from the MERGED review workbook
(REVISED_commands_merged_with_raw.xlsx), the single source of truth.
18 cols = Code | Sub Function | Test Set | Items | Procedure | Criteria |
          Bundle | Branch | Script | Attended Time | Machine Time |
          Latest Updated | ai_can_execute | ai_packages_needed |
          ai_commands | ai_logs_output | risk | remark

Every sheet becomes one library sheet. Each case carries the raw fields,
the AI-review fields, and extra fields (script/bundle/branch/attended_time/
machine_time/latest_updated/remark) when present.

Pure-ASCII script: any CJK in the JSON is carried over verbatim from the xlsx
via openpyxl (json dumped with ensure_ascii=False). No CJK literal here.

Usage: python3 scripts/build_testlib_json_xlsx.py [merged.xlsx] [out.json]
"""
import openpyxl, json, os, sys

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_WS = os.path.join(REPO, "data", "REVISED_commands_merged_with_raw.xlsx")
DEFAULT_OUT = "/srv/pa-manager-prod/data/tests.json"

SHEET_LABELS = {
    "Functionality": "\u529f\u80fd\u6027",          # Functional
    "Performance":   "\u6548\u80fd",                # Performance
    "Reliability":   "\u53ef\u9760\u6027",          # Reliability
    "Compatibility": "\u76f8\u5bb9\u6027",          # Compatibility
    "Stability":     "\u7a69\u5b9a\u6027",          # Stability
    "(No Main Function)": "\u7121\u4e3b\u529f\u80fd",  # No Main Function
}

FIELD_MAP = {
    "Code": "code",
    "Sub Function": "sub_function",
    "Test Set": "test_set",
    "Items": "items",
    "Procedure": "procedure",
    "Criteria": "criteria",
    "ai_can_execute": "ai_can_execute",
    "ai_packages_needed": "ai_packages_needed",
    "ai_commands": "ai_commands",
    "ai_logs_output": "ai_logs_output",
    "risk": "risk",
    "Script": "script",
    "Bundle": "bundle",
    "Branch": "branch",
    "Attended Time": "attended_time",
    "Machine Time": "machine_time",
    "Latest Updated": "latest_updated",
    "remark": "remark",
}


def main():
    ws_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WS
    out_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT
    wb = openpyxl.load_workbook(ws_path, read_only=True)

    library = {"sheets": {}, "total": 0, "generated_at": None}
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        hdr_raw = next(rows)
        col = {str(h).strip(): i for i, h in enumerate(hdr_raw) if h is not None}
        items = []
        for r in rows:
            code = r[col["Code"]] if "Code" in col else None
            if not code:
                continue
            it = {}
            for hname, jkey in FIELD_MAP.items():
                if hname not in col:
                    continue
                v = r[col[hname]]
                if v is None:
                    continue
                s = str(v).strip()
                if s == "":
                    continue
                it[jkey] = s
            it.setdefault("code", str(code).strip())
            items.append(it)
        if items:
            label = SHEET_LABELS.get(sn, sn)
            library["sheets"][label] = {
                "name": sn, "label": label,
                "count": len(items), "items": items,
            }
            library["total"] += len(items)

    # anti-template sanity
    import collections
    counter = collections.Counter()
    for s in library["sheets"].values():
        for it in s.get("items", []):
            c = str(it.get("ai_commands") or "").strip()
            if c:
                counter[c] += 1
    dups = {k: v for k, v in counter.items() if v > 1}
    if dups:
        print("WARNING: %d distinct command texts shared by >=2 cases (first 5):" % len(dups))
        for k, v in list(dups.items())[:5]:
            print("   x%d: %r" % (v, k[:120]))

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(library, f, ensure_ascii=False, indent=1)
    print("wrote", out_path, "total", library["total"])
    for label, s in library["sheets"].items():
        print("  ", label, s["name"], s["count"])


if __name__ == "__main__":
    main()
