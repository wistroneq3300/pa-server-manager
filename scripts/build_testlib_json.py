#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build tests.json (test library) by merging the lazy AI-review Excel with the
raw source Excel. Pure-ASCII script: any CJK in output JSON is carried over
verbatim from the source workbooks via openpyxl (using json ensure_ascii=True would
escape CJK, but we write UTF-8 directly so the review text is NOT re-typed). This script does NOT
any CJK literal. All UI-ish labels are emitted as unicode-escape (4-hex)
sequences so this file stays 7-bit clean (avoids LLM-output CJK corruption).

Output: /srv/pa-manager-prod/data/tests.json
"""
import openpyxl, json, os, sys, csv

RAW     = "/root/test-library/TestCaseLibrary_Wistron.xlsx"
REVIEW  = "/root/test-library/AI_Simplified_Review_20260828/TestCaseLibrary_Wistron_Simplified_AI_Review_20260828.xlsx"
OUT     = "/srv/pa-manager-prod/data/tests.json"

# T4 re-review sidecar + user additions (keyed by code; do NOT touch the workbooks)
# Prefer the repo-copy (data/) so the build is reproducible from a clone; fall
# back to /root/test-library/ for backward compatibility with this machine.
_REPO = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
_LEGACY = "/root/test-library"

def _pick(*names):
    repo_p = os.path.join(_REPO, names[0])
    if os.path.exists(repo_p):
        return repo_p
    return os.path.join(_LEGACY, names[0])

REVISED = _pick("REVISED_commands.csv")
ADDITIONS = _pick("ADDITIONS.csv")

# anti-template check
STRICT_NO_TEMPLATE = True

SHEET_LABELS = {
    "Functionality": "\u529f\u80fd\u6027",        # Functional
    "Performance":   "\u6548\u80fd",              # Performance
    "Reliability":   "\u53ef\u9760\u6027",        # Reliability
    "Compatibility": "\u76f8\u5bb9\u6027",        # Compatibility
    "Stability":     "\u7a69\u5b9a\u6027",        # Stability
    "(No Main Function)": "\u7121\u4e3b\u529f\u80fd",  # No Main Function
}

def col_map(hdr):
    return {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}

REVISED_FIELDS = ("ai_can_execute", "ai_commands", "ai_packages_needed",
                  "ai_logs_output", "risk")

def load_overlay(path):
    """Load a sidecar CSV (REVISED_commands.csv / ADDITIONS.csv) into
    {code: {field: value}}. Empty/missing file -> {}."""
    out = {}
    if not path or not os.path.exists(path):
        return out
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rd = csv.DictReader(f)
        for row in rd:
            code = (row.get("code") or "").strip()
            if not code:
                continue
            rec = {}
            for k in REVISED_FIELDS:
                rec[k] = (row.get(k) or "").strip()
            out[code] = rec
    return out

def overlay_item(item, revised):
    """Apply sidecar overlay (by code) onto an item dict. Keeps raw fields except
    the T4-reviewed ones + risk. Returns the (possibly new) item."""
    code = str(item.get("code", "")).strip()
    rec = revised.get(code)
    if rec:
        fields = list(rec)
        for k in fields:
            v = rec[k]
            if k == "risk" and v == "":
                item.pop("risk", None)
                continue
            if v == "":
                item.pop(k, None)
            else:
                item[k] = v
    elif "risk" not in item:
        item["risk"] = None
    return item

def main():
    wb_raw = openpyxl.load_workbook(RAW, read_only=True)
    wb_rev = openpyxl.load_workbook(REVIEW, read_only=True)

    raw_by_key = {}   # (sheet, code) -> row dict
    for sn in wb_raw.sheetnames:
        if sn == "Summary":
            continue
        idx = None
        for r in wb_raw[sn].iter_rows(values_only=True):
            if idx is None:
                idx = col_map(r)
                continue
            code = r[idx["Code"]] if "Code" in idx else None
            if not code:
                continue
            raw_by_key[(sn, str(code).strip())] = {
                "sub_function": r[idx.get("Sub Function", -1)] if "Sub Function" in idx else None,
                "test_set":     r[idx.get("Test Set", -1)] if "Test Set" in idx else None,
                "items":        r[idx.get("Items", -1)] if "Items" in idx else None,
                "procedure":    r[idx.get("Procedure", -1)] if "Procedure" in idx else None,
                "criteria":     r[idx.get("Criteria", -1)] if "Criteria" in idx else None,
            }

    wb_raw.close()

    revised = load_overlay(REVISED)

    library = {"sheets": {}, "total": 0, "generated_at": None}
    for sn in wb_rev.sheetnames:
        if sn == "Summary":
            continue
        idx = None
        items = []
        for r in wb_rev[sn].iter_rows(values_only=True):
            if idx is None:
                idx = col_map(r)
                continue
            code = r[idx["Code"]] if "Code" in idx else None
            if not code:
                continue
            key = (sn, str(code).strip())
            raw = raw_by_key.get(key) or {}
            items.append({
                "code":       str(code).strip(),
                "sub_function": raw.get("sub_function"),
                "test_set":   raw.get("test_set") or (r[idx["Test Set"]] if "Test Set" in idx else None),
                "items":      raw.get("items") or (r[idx["Items"]] if "Items" in idx else None),
                "procedure":  raw.get("procedure"),
                "criteria":   raw.get("criteria"),
                "ai_can_execute": r[idx["AI_Can_Execute"]] if "AI_Can_Execute" in idx else None,
                "ai_packages_needed": r[idx["AI_Packages_Needed"]] if "AI_Packages_Needed" in idx else None,
                "ai_commands": r[idx["AI_Commands"]] if "AI_Commands" in idx else None,
                "ai_logs_output": r[idx["AI_Logs_Output"]] if "AI_Logs_Output" in idx else None,
            })
            overlay_item(items[-1], revised)
            items[-1].setdefault("risk", None)
        if items:
            library["sheets"][SHEET_LABELS.get(sn, sn)] = {
                "name": sn,
                "label": SHEET_LABELS.get(sn, sn),
                "count": len(items),
                "items": items,
            }
            library["total"] += len(items)

    # Merge user's ADDITIONS.csv (new test cases) into the "Functionality" sheet.
    additions = load_overlay(ADDITIONS)
    if additions:
        # ADDITIONS rows carry code + all T4 fields; derive raw fields from
        # sub_function/test_set/items columns if present.
        add_rows = []
        with open(ADDITIONS, "r", encoding="utf-8-sig", newline="") as af:
            for row in csv.DictReader(af):
                code = (row.get("code") or "").strip()
                if not code:
                    continue
                ai = {}
                for k in REVISED_FIELDS:
                    ai[k] = (row.get(k) or "").strip()
                add_rows.append({
                    "code": code,
                    "sub_function": (row.get("sub_function") or "").strip() or None,
                    "test_set": (row.get("test_set") or "").strip() or None,
                    "items": (row.get("items") or "").strip() or None,
                    "procedure": (row.get("procedure") or "").strip() or None,
                    "criteria": (row.get("criteria") or "").strip() or None,
                    "ai_can_execute": ai["ai_can_execute"],
                    "ai_packages_needed": ai["ai_packages_needed"],
                    "ai_commands": ai["ai_commands"],
                    "ai_logs_output": ai["ai_logs_output"],
                    "risk": ai["risk"] or None,
                })
        if add_rows:
            fn_sheet = library["sheets"].setdefault(
                SHEET_LABELS["Functionality"], {
                    "name": "Functionality",
                    "label": SHEET_LABELS["Functionality"],
                    "count": 0,
                    "items": [],
                })
            seen = {it["code"] for it in fn_sheet["items"]}
            added = 0
            for a in add_rows:
                if a["code"] in seen:
                    continue
                fn_sheet["items"].append(a)
                seen.add(a["code"])
                library["total"] += 1
                added += 1
            fn_sheet["count"] = len(fn_sheet["items"])
            if added:
                print("ADDITIONS merged:", added, "new cases")

    # --- anti-template sanity: any two cases with byte-identical commands? ---
    if STRICT_NO_TEMPLATE:
        import collections
        counter = collections.Counter()
        for s in library["sheets"].values():
            for it in s.get("items", []):
                c = str(it.get("ai_commands") or "").strip()
                if c:
                    counter[c] += 1
        dups = {k: v for k, v in counter.items() if v > 1}
        if dups:
            print("WARNING: %d distinct command texts shared by >=2 cases (first shown):"
                  % len(dups))
            for k, v in list(dups.items())[:5]:
                print("   x%d: %r" % (v, k[:120]))

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(library, f, ensure_ascii=False, indent=1)
    print("wrote", OUT, "total", library["total"])
    for label, s in library["sheets"].items():
        print("  ", label, s["name"], s["count"])

if __name__ == "__main__":
    main()
